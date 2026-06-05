import json
import signal
import sys
import time
from datetime import datetime

import click
from loguru import logger
from rich.console import Console
from rich.table import Table
from sqlalchemy.orm import sessionmaker

from config import DEFAULT_CITIES, DEFAULT_SECTORS, MAX_PROSPECTS_PER_RUN, LOG_FILE, EXPLORIUM_API_KEY
from database.models import get_engine, init_db, ProspectStatus
from database.crud import (
    create_prospect, get_prospects, get_stats,
    prospect_exists, count_emails_sent_today
)
from prospecting.explorium_client import ExploRiumClient
from prospecting.website_analyzer import WebsiteAnalyzer
from email_agent.email_writer import generate_email_templates
from email_agent.email_templates import (
    init_template_db, get_or_create_templates, update_template,
    get_template, list_all_templates
)
from email_agent.sender import send_prospecting_email
from email_agent.gmail_client import GmailClient
from scheduler.follow_up_scheduler import FollowUpScheduler

logger.add(LOG_FILE, rotation="10 MB", retention="30 days", level="INFO")
console = Console()


def get_db():
    engine = get_engine()
    SessionLocal = sessionmaker(bind=engine)
    return SessionLocal()


@click.group()
def cli():
    pass


@cli.command()
@click.option("--sectors", default=None, help="Secteurs séparés par des virgules")
@click.option("--max-results", default=MAX_PROSPECTS_PER_RUN, type=int)
@click.option("--dry-run", is_flag=True, help="Affiche sans sauvegarder")
@click.option("--demo", is_flag=True, help="Mode démo avec données fictives")
def prospect(sectors, max_results, dry_run, demo):
    """Recherche de nouveaux prospects via Explorium (ou démo)."""
    init_db()
    db = get_db()

    if demo:
        from prospecting.demo_data import generate_demo_prospects
        console.print("[bold yellow][DEMO MODE][/bold yellow]")
        console.print("Utilisation de données fictives pour tester\n")
        results = generate_demo_prospects(max_results)
    else:
        if not EXPLORIUM_API_KEY:
            console.print("[red][X][/red] EXPLORIUM_API_KEY manquante dans .env")
            console.print("Conseil: utilise [bold]--demo[/bold] pour tester sans crédit")
            return

        sector_list = [s.strip() for s in sectors.split(",")] if sectors else DEFAULT_SECTORS

        console.print(f"[bold blue]Recherche Explorium en cours...[/bold blue]")
        console.print(f"Secteurs: {sector_list}")

        client = ExploRiumClient()

        results = client.run_full_search(
            sectors=sector_list,
            countries=["fr"],
            max_per_sector=max_results // len(sector_list) if sector_list else max_results
        )

    analyzer = WebsiteAnalyzer()

    console.print(f"[green]{len(results)} entreprises trouvées[/green]")

    saved = 0
    skipped = 0

    for business in results:
        company_name = business.get("company_name", "").strip()
        city = business.get("city", "").strip()
        domain = business.get("domain", "")
        phone = business.get("phone")

        if not company_name:
            continue

        if not dry_run and prospect_exists(db, company_name, city):
            skipped += 1
            continue

        if "website_score" in business:
            analysis = {"score": business.get("website_score"), "issues": business.get("website_issues", []), "cms": None}
        else:
            analysis = analyzer.analyze(domain) if domain else {"score": 0, "issues": ["Aucun site web"], "cms": None}

        if not analyzer.is_worth_contacting(analysis["score"]):
            skipped += 1
            continue

        prospects_list = business.get("prospects", [])
        email = business.get("email") or (prospects_list[0].get("email") if prospects_list else "")

        prospect_data = {
            "company_name": company_name,
            "industry": business.get("industry", ""),
            "address": business.get("address", ""),
            "city": city,
            "phone": phone or "",
            "email": email or "",
            "website_url": domain or "",
            "website_score": analysis["score"],
            "website_issues": json.dumps(business.get("website_issues") or analysis["issues"], ensure_ascii=False),
            "source": "demo" if demo else "explorium",
            "status": ProspectStatus.new
        }

        if dry_run:
            console.print(f"[cyan][DRY RUN][/cyan] {company_name} ({city}) — score: {analysis['score']} — email: {email or 'non trouvé'}")
        else:
            create_prospect(db, prospect_data)
            saved += 1

    db.close()
    console.print(f"[bold green]Terminé![/bold green] {saved} sauvegardés, {skipped} ignorés.")


@cli.command("validate-templates")
@click.option("--demo", is_flag=True, help="Utiliser les templates de démo (sans API Claude)")
def validate_templates(demo):
    """Générer et valider les 2 templates d'email (une seule fois)."""
    init_template_db()
    db = get_db()

    if demo:
        from email_agent.demo_templates import DEMO_TEMPLATES
        console.print("[bold yellow][TEMPLATES DE DEMO][/bold yellow]")
        console.print("Utilisation des templates prédéfinis pour tester\n")
        templates = DEMO_TEMPLATES
    else:
        console.print("[bold blue]Génération des templates par Claude...[/bold blue]\n")
        templates = generate_email_templates()

    no_website = templates.get("no_website", {})
    bad_website = templates.get("bad_website", {})

    console.print("[bold green]=== TEMPLATE 1 : Entreprises SANS SITE WEB ===[/bold green]")
    console.print(f"Objet: {no_website.get('subject', '')}\n")
    console.print(f"Corps:\n{no_website.get('body', '')}\n")
    console.print("[yellow]" + "-" * 60 + "[/yellow]\n")

    console.print("[bold green]=== TEMPLATE 2 : Entreprises AVEC SITE MÉDIOCRE ===[/bold green]")
    console.print(f"Objet: {bad_website.get('subject', '')}\n")
    console.print(f"Corps:\n{bad_website.get('body', '')}\n")
    console.print("[yellow]" + "-" * 60 + "[/yellow]\n")

    if demo or click.confirm("Validez-vous ces 2 templates ?"):
        get_or_create_templates(db, "no_website", no_website.get("subject", ""), no_website.get("body", ""))
        get_or_create_templates(db, "bad_website", bad_website.get("subject", ""), bad_website.get("body", ""))
        console.print("[green][OK] Templates sauvegardés et validés![/green]")
    else:
        console.print("[yellow]Validation annulée. Relancez cette commande pour réessayer.[/yellow]")

    db.close()


@cli.command("send-emails")
@click.option("--limit", default=20, type=int, help="Nb max d'emails à envoyer")
@click.option("--dry-run", is_flag=True)
@click.option("--status", default="new")
def send_emails(limit, dry_run, status):
    """Envoie les emails aux prospects."""
    init_db()
    db = get_db()

    try:
        s = ProspectStatus(status)
    except ValueError:
        console.print(f"[red]Statut invalide: {status}[/red]")
        return

    prospects = get_prospects(db, status=s, limit=limit)
    console.print(f"[blue]{len(prospects)} prospects à contacter[/blue]")

    sent = 0
    for p in prospects:
        if not p.email:
            console.print(f"[yellow]Pas d'email: {p.company_name}[/yellow]")
            continue

        if dry_run:
            console.print(f"[cyan][DRY RUN][/cyan] Enverrait à {p.company_name} <{p.email}>")
            console.print(f"  Score site: {p.website_score} | Issues: {p.website_issues}")
        else:
            if send_prospecting_email(db, p.id):
                sent += 1
                console.print(f"[green]✓[/green] {p.company_name} ({p.email})")
            else:
                console.print(f"[red]✗[/red] {p.company_name}")

    db.close()
    if not dry_run:
        console.print(f"[bold green]{sent} emails envoyés[/bold green]")


@cli.command("start-agent")
def start_agent():
    """Lance l'agent en mode daemon."""
    init_db()
    console.print("[bold green]Démarrage de l'agent...[/bold green]")

    scheduler = FollowUpScheduler()
    scheduler.start()

    console.print("[green]Agent démarré! Jobs planifiés:[/green]")
    for job in scheduler.get_jobs():
        console.print(f"  • {job.name}: prochain à {job.next_run_time}")

    console.print("\n[yellow]Appuyez sur Ctrl+C pour arrêter[/yellow]")

    def handle_stop(sig, frame):
        console.print("\n[yellow]Arrêt de l'agent...[/yellow]")
        scheduler.stop()
        sys.exit(0)

    signal.signal(signal.SIGINT, handle_stop)
    signal.signal(signal.SIGTERM, handle_stop)

    while True:
        time.sleep(60)


@cli.command()
def status():
    """Affiche un résumé des prospects."""
    init_db()
    db = get_db()

    stats = get_stats(db)
    sent_today = count_emails_sent_today(db)

    table = Table(title="Statut de l'agent de prospection")
    table.add_column("Métrique", style="cyan")
    table.add_column("Valeur", style="green")

    table.add_row("Total prospects", str(stats["total"]))
    for s, count in stats["by_status"].items():
        table.add_row(f"  └ {s}", str(count))
    table.add_row("Emails envoyés aujourd'hui", str(sent_today))

    console.print(table)

    from database.crud import get_prospects_for_followup
    followups = get_prospects_for_followup(db)
    if followups:
        console.print(f"\n[yellow]{len(followups)} relances en attente[/yellow]")

    db.close()


@cli.command("list-prospects")
@click.option("--status", default=None)
@click.option("--city", default=None)
@click.option("--score-max", default=None, type=int)
@click.option("--limit", default=50, type=int)
def list_prospects(status, city, score_max, limit):
    """Affiche la liste des prospects."""
    init_db()
    db = get_db()

    s = None
    if status:
        try:
            s = ProspectStatus(status)
        except ValueError:
            console.print(f"[red]Statut invalide: {status}[/red]")
            return

    prospects = get_prospects(db, status=s, city=city, score_max=score_max, limit=limit)

    table = Table(title=f"Prospects ({len(prospects)})")
    table.add_column("ID", style="dim")
    table.add_column("Entreprise")
    table.add_column("Secteur")
    table.add_column("Ville")
    table.add_column("Email")
    table.add_column("Score", justify="right")
    table.add_column("Statut")

    for p in prospects:
        score_color = "red" if p.website_score <= 40 else "yellow" if p.website_score <= 65 else "green"
        table.add_row(
            str(p.id),
            p.company_name,
            p.industry or "",
            p.city or "",
            p.email or "[dim]non trouvé[/dim]",
            f"[{score_color}]{p.website_score}[/{score_color}]",
            p.status.value if p.status else ""
        )

    console.print(table)
    db.close()


@cli.command("export-prospects")
@click.option("--format", default="excel", type=click.Choice(["excel", "csv"]), help="Format export (excel ou csv)")
@click.option("--status", default=None, help="Filtrer par statut")
@click.option("--output", default=None, help="Nom du fichier (défaut: prospects_YYYY-MM-DD.xlsx)")
def export_prospects(format, status, output):
    """Exporte les prospects en Excel ou CSV."""
    init_db()
    db = get_db()

    s = None
    if status:
        try:
            s = ProspectStatus(status)
        except ValueError:
            console.print(f"[red][X][/red] Statut invalide: {status}")
            return

    prospects = get_prospects(db, status=s, limit=10000)

    if not prospects:
        console.print("[yellow]Aucun prospect à exporter[/yellow]")
        return

    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d")
    default_filename = f"prospects_{timestamp}"

    if format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            filename = output or f"{default_filename}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Prospects"

            headers = ["ID", "Entreprise", "Contact", "Email", "Téléphone", "Secteur", "Adresse", "Ville", "Score Site", "Statut", "Date Création"]
            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for prospect in prospects:
                ws.append([
                    prospect.id,
                    prospect.company_name,
                    prospect.company_name,
                    prospect.email or "",
                    prospect.phone or "",
                    prospect.industry or "",
                    prospect.address or "",
                    prospect.city or "",
                    prospect.website_score or 0,
                    prospect.status.value if prospect.status else "",
                    prospect.created_at.strftime("%d/%m/%Y") if prospect.created_at else ""
                ])

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filename)
            console.print(f"[green][OK][/green] {len(prospects)} prospects exportés dans [bold]{filename}[/bold]")

        except ImportError:
            console.print("[red][X][/red] openpyxl non installé. Installe: pip install openpyxl")

    else:
        import csv
        filename = output or f"{default_filename}.csv"

        with open(filename, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = ["ID", "Entreprise", "Contact", "Email", "Téléphone", "Secteur", "Adresse", "Ville", "Score Site", "Statut", "Date Création"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for prospect in prospects:
                writer.writerow({
                    "ID": prospect.id,
                    "Entreprise": prospect.company_name,
                    "Contact": prospect.company_name,
                    "Email": prospect.email or "",
                    "Téléphone": prospect.phone or "",
                    "Secteur": prospect.industry or "",
                    "Adresse": prospect.address or "",
                    "Ville": prospect.city or "",
                    "Score Site": prospect.website_score or 0,
                    "Statut": prospect.status.value if prospect.status else "",
                    "Date Création": prospect.created_at.strftime("%d/%m/%Y") if prospect.created_at else ""
                })

        console.print(f"[green][OK][/green] {len(prospects)} prospects exportés dans [bold]{filename}[/bold]")

    db.close()


@cli.command()
def setup():
    """Guide d'installation interactif."""
    console.print("[bold blue]=== Configuration de l'agent de prospection ===[/bold blue]\n")

    console.print("[bold]Étape 1: Vérification des clés API[/bold]")
    from config import ANTHROPIC_API_KEY, GMAIL_SENDER_EMAIL

    if ANTHROPIC_API_KEY:
        console.print("[green][OK][/green] ANTHROPIC_API_KEY configurée")
    else:
        console.print("[red][X][/red] ANTHROPIC_API_KEY manquante dans .env")

    if EXPLORIUM_API_KEY:
        console.print("[green][OK][/green] EXPLORIUM_API_KEY configurée")
    else:
        console.print("[red][X][/red] EXPLORIUM_API_KEY manquante dans .env")

    if GMAIL_SENDER_EMAIL:
        console.print(f"[green][OK][/green] Email expéditeur: {GMAIL_SENDER_EMAIL}")
    else:
        console.print("[red][X][/red] GMAIL_SENDER_EMAIL manquant dans .env")

    console.print("\n[bold]Étape 2: Authentification Gmail OAuth2[/bold]")
    if click.confirm("Lancer l'authentification Gmail maintenant ?"):
        try:
            gmail = GmailClient()
            gmail.authenticate()
            console.print("[green][OK][/green] Gmail authentifié avec succès!")
        except Exception as e:
            console.print(f"[red]✗[/red] Erreur: {e}")
            console.print("Vérifiez que credentials.json est présent dans le répertoire.")

    console.print("\n[bold]Étape 3: Initialisation de la base de données[/bold]")
    try:
        init_db()
        console.print("[green][OK][/green] Base de données initialisée (prospects.db)")
    except Exception as e:
        console.print(f"[red]✗[/red] Erreur BDD: {e}")

    console.print("\n[bold]Étape 4: Test d'envoi d'email[/bold]")
    from config import NOTIFICATION_EMAIL
    if NOTIFICATION_EMAIL and click.confirm(f"Envoyer un email de test à {NOTIFICATION_EMAIL} ?"):
        try:
            gmail = GmailClient()
            gmail.send_email(
                to=NOTIFICATION_EMAIL,
                subject="[TEST] Agent de prospection opérationnel",
                body_html="<p>Votre agent de prospection est configuré et opérationnel !</p>",
                body_text="Votre agent de prospection est configuré et opérationnel !"
            )
            console.print("[green][OK][/green] Email de test envoyé!")
        except Exception as e:
            console.print(f"[red]✗[/red] Erreur envoi test: {e}")

    console.print("\n[bold green]Configuration terminée![/bold green]")
    console.print("Commandes suivantes:")
    console.print("  python main.py validate-templates  — Générer et valider les 2 templates (UNE FOIS)")
    console.print("  python main.py prospect            — Recherche de nouveaux prospects")
    console.print("  python main.py send-emails         — Envoyer les emails (déploiement auto)")
    console.print("  python main.py start-agent         — Lance l'agent automatique 24/7")
    console.print("  python main.py status              — Tableau de bord")


if __name__ == "__main__":
    cli()
