from flask import Flask, jsonify, send_from_directory, request, send_file
from flask_cors import CORS
from sqlalchemy.orm import sessionmaker
import json
import os
from datetime import datetime
import traceback

from database.models import get_engine, init_db, ProspectStatus, EmailLog, EmailDirection
from database.crud import (
    get_prospects, get_stats, create_prospect, get_prospect,
    count_emails_sent_today, get_prospects_for_followup, create_email_log
)
from prospecting.finder import ProspectFinder
from prospecting.website_analyzer import WebsiteAnalyzer
from prospecting.verifier import verify_web_presence, get_verification, set_verification
from email_agent.gmail_client import GmailClient
from config import TEST_MODE

app = Flask(__name__, static_folder='static', static_url_path='/static')
CORS(app)
init_db()

def get_db():
    engine = get_engine()
    SessionLocal = sessionmaker(bind=engine)
    return SessionLocal()

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    if path != '' and os.path.exists(os.path.join('static', path)):
        return send_from_directory('static', path)
    return send_from_directory('static', 'index.html')

@app.errorhandler(404)
def not_found(e):
    return send_from_directory('static', 'index.html')

@app.route('/api/stats')
def get_stats_api():
    try:
        db = get_db()
        stats = get_stats(db)

        total = stats.get('total', 0)
        by_status = stats.get('by_status', {})

        result = {
            'total': total,
            'emails_sent': count_emails_sent_today(db),
            'replied': by_status.get('replied', 0),
            'contacted': by_status.get('contacted', 0),
            'new': by_status.get('new', 0),
            'converted': by_status.get('converted', 0),
            'pending_followups': len(get_prospects_for_followup(db))
        }
        db.close()
        return jsonify(result)
    except Exception as e:
        print(f"Erreur /api/stats: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/prospects')
def get_prospects_api():
    try:
        db = get_db()

        status = request.args.get('status')
        city = request.args.get('city')
        limit = int(request.args.get('limit', 100))

        status_obj = None
        if status and status != '' and status != 'all':
            try:
                status_obj = ProspectStatus(status)
            except ValueError:
                status_obj = None

        prospects = get_prospects(db, status=status_obj, city=city if city else None, limit=limit)

        result = []
        for p in prospects:
            result.append({
                'id': p.id,
                'company_name': p.company_name,
                'industry': p.industry or '',
                'city': p.city or '',
                'email': p.email or '',
                'phone': p.phone or '',
                'address': p.address or '',
                'website_score': p.website_score or 0,
                'website_url': p.website_url or '',
                'status': p.status.value if p.status else 'unknown',
                'created_at': p.created_at.isoformat() if p.created_at else '',
                'notes': p.notes or '',
                'contact_count': p.contact_count or 0,
            })

        db.close()
        return jsonify({'prospects': result, 'count': len(result)})
    except Exception as e:
        print(f"Erreur /api/prospects: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/prospect/<int:prospect_id>')
def get_prospect_api(prospect_id):
    try:
        db = get_db()
        prospect = get_prospect(db, prospect_id)

        if not prospect:
            db.close()
            return jsonify({'error': 'Prospect not found'}), 404

        result = {
            'id': prospect.id,
            'company_name': prospect.company_name,
            'industry': prospect.industry,
            'city': prospect.city,
            'email': prospect.email,
            'phone': prospect.phone,
            'address': prospect.address,
            'website_score': prospect.website_score,
            'website_url': prospect.website_url,
            'status': prospect.status.value if prospect.status else 'unknown'
        }
        db.close()
        return jsonify(result)
    except Exception as e:
        print(f"Erreur /api/prospect/{prospect_id}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/search', methods=['POST'])
def search_api():
    try:
        data = request.json
        sector = data.get('sector', '').strip()
        city = data.get('city', '').strip()
        max_results = min(int(data.get('max_results', 20)), 100)
        no_website_only = bool(data.get('no_website_only', False))

        if not sector or not city:
            return jsonify({'success': False, 'error': 'Secteur et ville requis'}), 400

        print(f"[SEARCH] {sector} à {city} (max={max_results}, sans_site={no_website_only})")

        # max_score: 0 = sans site uniquement, 65 = sans site + site médiocre
        max_score = 0 if no_website_only else int(data.get('max_score', 65))

        finder = ProspectFinder()

        results = finder.run_full_search(
            sectors=[sector],
            cities=[city],
            max_per_combo=max_results,
            no_website_only=no_website_only,
            max_score=max_score,
        )

        print(f"[SEARCH] {len(results)} prospects qualifiés, sauvegarde en cours...")

        db = get_db()
        existing_names = {p.company_name.lower().strip() for p in get_prospects(db, limit=50000)}

        saved = 0
        for business in results:
            company_name = business.get('name', '').strip()
            city_name = business.get('city', city).strip()
            if not company_name or company_name.lower() in existing_names:
                continue

            domain = business.get('website', '')
            # Use pre-computed score from finder (avoids double analysis)
            ws_score = business.get('website_score', 0)
            ws_issues = business.get('website_issues', ['Aucun site web'] if not domain else [])
            ws_label = business.get('website_label', '')
            ws_positives = business.get('website_positives', [])
            ws_cms = business.get('website_cms', '')

            # Store rich analysis in issues field as JSON
            issues_payload = {
                'issues': ws_issues,
                'label': ws_label,
                'positives': ws_positives,
                'cms': ws_cms,
                'load_time': business.get('load_time'),
            }

            prospect_data = {
                "company_name": company_name,
                "industry": business.get('industry', sector),
                "address": business.get('address', ''),
                "city": city_name,
                "phone": business.get('phone', ''),
                "email": business.get('email', ''),
                "website_url": domain or "",
                "website_score": ws_score,
                "website_issues": json.dumps(issues_payload, ensure_ascii=False),
                "source": business.get('source', 'search'),
                "status": ProspectStatus.new
            }

            create_prospect(db, prospect_data)
            existing_names.add(company_name.lower())
            saved += 1

        db.commit()
        db.close()
        print(f"[SEARCH] {saved} prospects sauvegardés")
        return jsonify({
            'success': True,
            'saved': saved,
            'total_found': len(results),
            'message': f"{saved} prospects trouvés pour '{sector}' à {city}"
        })

    except Exception as e:
        print(f"Erreur /api/search: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/send-email', methods=['POST'])
def send_email_api():
    db = None
    try:
        data = request.json
        prospect_id = data.get('prospect_id')
        subject = data.get('subject')
        body = data.get('body')

        db = get_db()
        prospect = get_prospect(db, prospect_id)

        if not prospect:
            return jsonify({'success': False, 'error': 'Prospect not found'}), 404

        if not prospect.email:
            return jsonify({'success': False, 'error': 'No email for this prospect'}), 400

        prospect_email = prospect.email
        gmail_message_id = None

        if TEST_MODE:
            print(f"[TEST MODE] Email à {prospect_email}")
            print(f"Objet: {subject}")
            print(f"Corps: {body[:100]}...")
            gmail_message_id = f'test_mode_{prospect_id}_{datetime.now().timestamp()}'
        else:
            try:
                gmail_client = GmailClient()
                gmail_client.authenticate()
                result = gmail_client.send_email(
                    to=prospect_email,
                    subject=subject,
                    body_html=body,
                    body_text=body
                )
                gmail_message_id = result.get('message_id', f'email_{prospect_id}')
                print(f"Email envoyé à {prospect_email} (ID: {gmail_message_id})")
            except Exception as e:
                print(f"Erreur envoi Gmail: {e}")
                gmail_message_id = f'error_{prospect_id}'
                if db:
                    db.close()
                return jsonify({'success': False, 'error': f'Erreur envoi email: {str(e)}'}), 500

        create_email_log(db, {
            'prospect_id': prospect_id,
            'direction': EmailDirection.sent,
            'subject': subject,
            'body': body,
            'sent_at': datetime.now(),
            'gmail_message_id': gmail_message_id
        })

        db.commit()

        mode_text = " (TEST MODE - non envoyé)" if TEST_MODE else ""
        return jsonify({'success': True, 'message': f'Email enregistré pour {prospect_email}{mode_text}'})

    except Exception as e:
        print(f"Erreur /api/send-email: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if db:
            db.close()

@app.route('/api/emails')
def get_emails_api():
    try:
        db = get_db()

        direction = request.args.get('direction', 'sent')
        limit = int(request.args.get('limit', 100))

        query = db.query(EmailLog)
        if direction != 'all':
            query = query.filter(EmailLog.direction == direction)

        emails = query.order_by(EmailLog.sent_at.desc()).limit(limit).all()

        result = []
        for e in emails:
            prospect = get_prospect(db, e.prospect_id)
            result.append({
                'id': e.id,
                'prospect_id': e.prospect_id,
                'company_name': prospect.company_name if prospect else 'Unknown',
                'industry': prospect.industry if prospect else '',
                'email': prospect.email if prospect else '',
                'direction': e.direction.value if e.direction else 'unknown',
                'subject': e.subject,
                'body': e.body,
                'sent_at': e.sent_at.isoformat() if e.sent_at else ''
            })

        db.close()
        return jsonify({'emails': result, 'count': len(result)})

    except Exception as e:
        print(f"Erreur /api/emails: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/export')
def export_api():
    try:
        format_type = request.args.get('format', 'csv')

        db = get_db()
        prospects = get_prospects(db, limit=10000)

        data = []
        for p in prospects:
            data.append({
                'ID': p.id,
                'Entreprise': p.company_name,
                'Email': p.email or '',
                'Téléphone': p.phone or '',
                'Secteur': p.industry or '',
                'Adresse': p.address or '',
                'Ville': p.city or '',
                'Score Site': p.website_score or 0,
                'Statut': p.status.value if p.status else '',
                'Créé': p.created_at.strftime('%d/%m/%Y') if p.created_at else ''
            })

        if format_type == 'csv':
            import csv
            from io import StringIO

            output = StringIO()
            fieldnames = ['ID', 'Entreprise', 'Email', 'Téléphone', 'Secteur', 'Adresse', 'Ville', 'Score Site', 'Statut', 'Créé']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

            filename = f"prospects_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.csv"
            db.close()
            return output.getvalue(), 200, {'Content-Disposition': f'attachment; filename={filename}', 'Content-Type': 'text/csv; charset=utf-8'}

        else:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Prospects"

                headers = ['ID', 'Entreprise', 'Email', 'Téléphone', 'Secteur', 'Adresse', 'Ville', 'Score Site', 'Statut', 'Créé']
                ws.append(headers)

                header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                for row in data:
                    ws.append([row.get(h, '') for h in headers])

                for idx, col in enumerate(ws.columns, 1):
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 50)

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                filename = f"prospects_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
                db.close()
                return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

            except Exception as e:
                db.close()
                return jsonify({'error': str(e)}), 500

    except Exception as e:
        print(f"Erreur /api/export: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/send-test-email', methods=['POST'])
def send_test_email_api():
    db = None
    try:
        from config import GMAIL_SENDER_EMAIL
        data = request.json
        sector = data.get('sector', 'test')
        subject = data.get('subject', 'Email de test')
        body = data.get('body', 'Ceci est un email de test')

        test_email = GMAIL_SENDER_EMAIL

        gmail_message_id = None

        if TEST_MODE:
            print(f"[TEST MODE] Email de test à {test_email}")
            print(f"Secteur: {sector}")
            print(f"Objet: {subject}")
            gmail_message_id = f'test_email_{sector}_{datetime.now().timestamp()}'
        else:
            try:
                gmail_client = GmailClient()
                gmail_client.authenticate()
                result = gmail_client.send_email(
                    to=test_email,
                    subject=f"[TEST] {subject}",
                    body_html=body,
                    body_text=body
                )
                gmail_message_id = result.get('message_id', f'test_email_{sector}')
                print(f"Email de test envoyé à {test_email}")
            except Exception as e:
                print(f"Erreur envoi test: {e}")
                return jsonify({'success': False, 'error': f'Erreur: {str(e)}'}), 500

        db = get_db()
        create_email_log(db, {
            'prospect_id': None,
            'direction': EmailDirection.sent,
            'subject': f"[TEST - {sector}] {subject}",
            'body': body,
            'sent_at': datetime.now(),
            'gmail_message_id': gmail_message_id
        })
        db.commit()

        mode_text = " (TEST MODE)" if TEST_MODE else ""
        return jsonify({'success': True, 'message': f'Email de test enregistré{mode_text}'})

    except Exception as e:
        print(f"Erreur /api/send-test-email: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if db:
            db.close()

@app.route('/api/templates', methods=['GET'])
def get_templates_api():
    db = None
    try:
        from email_agent.email_templates import EmailTemplate
        db = get_db()
        templates = db.query(EmailTemplate).all()
        result = {}
        for t in templates:
            result[t.template_type] = {
                'subject': t.subject,
                'body': t.body
            }
        return jsonify({'templates': result})
    except Exception as e:
        print(f"Erreur /api/templates: {e}")
        print(traceback.format_exc())
        return jsonify({'templates': {}})
    finally:
        if db:
            db.close()

@app.route('/api/template/<sector>', methods=['POST'])
def save_template_api(sector):
    db = None
    try:
        from email_agent.email_templates import EmailTemplate
        data = request.json
        subject = data.get('subject', '')
        body = data.get('body', '')

        db = get_db()

        template = db.query(EmailTemplate).filter_by(template_type=sector).first()
        if template:
            template.subject = subject
            template.body = body
        else:
            template = EmailTemplate(template_type=sector, subject=subject, body=body)
            db.add(template)

        db.commit()
        return jsonify({'success': True, 'message': f'Template {sector} sauvegardé'})
    except Exception as e:
        print(f"Erreur /api/template/{sector}: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if db:
            db.close()

@app.route('/api/test')
def test_api():
    return jsonify({'status': 'ok', 'message': 'API is working'})


@app.route('/api/verify/<int:prospect_id>', methods=['POST'])
def verify_prospect(prospect_id):
    """Vérifie si un prospect a vraiment un site web."""
    db = None
    try:
        db = get_db()
        prospect = get_prospect(db, prospect_id)
        if not prospect:
            return jsonify({'success': False, 'error': 'Prospect not found'}), 404

        print(f"[VERIFY] {prospect.company_name} ({prospect.city})")
        result = verify_web_presence(
            company_name=prospect.company_name,
            city=prospect.city or '',
            existing_url=prospect.website_url or '',
        )

        # Persist result in notes
        prospect.notes = set_verification(prospect.notes, result)

        # If a real website was found, update website_url and re-score
        if result['has_website'] and result.get('found_url'):
            if not prospect.website_url:
                prospect.website_url = result['found_url']
            if prospect.website_score == 0:
                from prospecting.website_analyzer import WebsiteAnalyzer
                analysis = WebsiteAnalyzer().analyze(result['found_url'])
                prospect.website_score = analysis.get('score', 0)
                prospect.website_issues = json.dumps(analysis.get('issues', []), ensure_ascii=False)

        db.commit()
        print(f"[VERIFY] → has_website={result['has_website']}, confidence={result['confidence']}, reason={result['reason']}")
        return jsonify({'success': True, **result})

    except Exception as e:
        print(f"Erreur /api/verify/{prospect_id}: {e}")
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if db:
            db.close()


@app.route('/api/verify-bulk', methods=['POST'])
def verify_bulk():
    """Vérifie un lot de prospects (max 20 à la fois)."""
    db = None
    try:
        data = request.json or {}
        ids = data.get('ids', [])[:20]
        if not ids:
            # Verify all unverified prospects (status=new, no verification yet)
            db = get_db()
            prospects = get_prospects(db, status=ProspectStatus.new, limit=20)
            ids = [p.id for p in prospects
                   if get_verification(p.notes) is None]
            db.close()
            db = None

        results = {}
        db = get_db()
        for pid in ids:
            p = get_prospect(db, pid)
            if not p:
                continue
            try:
                result = verify_web_presence(p.company_name, p.city or '', p.website_url or '')
                p.notes = set_verification(p.notes, result)
                if result['has_website'] and result.get('found_url') and not p.website_url:
                    p.website_url = result['found_url']
                db.commit()
                results[pid] = result
            except Exception as e:
                results[pid] = {'error': str(e)}

        db.close()
        verified = sum(1 for r in results.values() if r.get('has_website') is False)
        has_site = sum(1 for r in results.values() if r.get('has_website') is True)
        return jsonify({
            'success': True,
            'checked': len(results),
            'no_website': verified,
            'has_website': has_site,
            'results': results,
        })

    except Exception as e:
        print(f"Erreur /api/verify-bulk: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if db:
            db.close()


@app.route('/api/prospect/<int:prospect_id>/verification')
def get_prospect_verification(prospect_id):
    """Returns the verification status for a prospect."""
    db = None
    try:
        db = get_db()
        prospect = get_prospect(db, prospect_id)
        if not prospect:
            return jsonify({'error': 'Not found'}), 404
        verification = get_verification(prospect.notes)
        return jsonify({'verification': verification})
    finally:
        if db:
            db.close()

if __name__ == '__main__':
    print("Démarrage du serveur sur http://127.0.0.1:8080")
    print("API endpoints disponibles:")
    print("  GET  /api/stats")
    print("  GET  /api/prospects")
    print("  GET  /api/prospect/<id>")
    print("  POST /api/search")
    print("  POST /api/send-email")
    print("  POST /api/send-test-email")
    print("  GET  /api/emails")
    print("  GET  /api/export")
    print("  GET  /api/test")
    app.run(debug=False, port=8080, host='127.0.0.1', use_reloader=False)
